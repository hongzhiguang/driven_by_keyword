from openpyxl import load_workbook
import traceback

class ParseExcel(object):

    def __init__(self,excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(self.excel_file_path)
        self.ws = self.set_sheet_by_name(self.wb.sheetnames[0])

    def set_sheet_by_name(self,name):
        # 设置操作的sheet
        if name in self.wb.sheetnames:
            self.ws = self.wb[name]
            return self.ws
        self.ws = None
        return self.ws

    def get_cols(self, row_no):
        if not isinstance(row_no,int):
            return None
        try:
            return list(self.ws.rows)[row_no-1]
        except:
            traceback.print_exc()

    def get_rows(self,col_no):
        if not isinstance(col_no,int):
            return None
        try:
            return list(self.ws.columns)[col_no-1]
        except:
            traceback.print_exc()

    def get_cell_value(self,row_no,col_no):
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            return None
        try:
            return self.ws.cell(row=row_no,column=col_no).value
        except:
            traceback.print_exc()

    def write_cell(self,row_no,col_no,value):
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            return None
        try:
            self.ws.cell(row=row_no,column=col_no).value = value
            self._save()
        except:
            traceback.print_exc()


    def get_max_row(self):
        try:
            return self.ws.max_row
        except:
            traceback.print_exc()
            return None

    def _save(self):
        # 表格中写入数据，保存生效
        self.wb.save(self.excel_file_path)



if __name__ == "__main__":
    pass
